import os

from pypdf import PdfReader
import zipfile
from openpyxl import load_workbook
from paths import RESOURCE_DIR
import csv


def test_csv():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, 'files.zip')) as zip_file:
        with zip_file.open('schedule.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            third_row = csvreader[2]

    assert third_row[0] == 'школа'
    assert third_row[5] == 'выходной'

def test_pdf():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, 'files.zip')) as zip_file:
        with zip_file.open('Колобок.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()

    assert text.find('Автор: Русская народная') != -1

def test_xlsx():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, 'files.zip')) as zip_file:
        with zip_file.open('pers.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

    assert sheet.cell(row=2, column=8).value == '88002000600'
